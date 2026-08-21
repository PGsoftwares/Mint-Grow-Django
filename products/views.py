from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from django.db.models import Q, Min, Max
from django.core.paginator import Paginator

from accounts.decorators import admin_required

from .forms import ProductForm, ProductPriceVariationFormSet
from .models import Product, ProductCategory

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date

def public_product_list_view(request):

    products = (
        Product.objects
        .select_related("category")
        .prefetch_related("price_variations")
        .filter(
            status="active",
            category__status="active",
        )
        .annotate(
            min_price=Min(
                "price_variations__price"
            ),
            max_price=Max(
                "price_variations__price"
            ),
        )
        .order_by(
            "-featured",
            "-created_at",
        )
    )


    categories = (
        ProductCategory.objects
        .filter(
            status="active",
            products__status="active",
        )
        .distinct()
        .order_by("name")
    )


    context = {
        "products": products,
        "categories": categories,
    }


    return render(
        request,
        "products/public_list.html",
        context,
    )


def public_product_detail_view(request, slug):
    product = get_object_or_404(
        Product.objects.select_related("category"),
        slug=slug,
        status="active",
        category__status="active",
    )

    related_products = (
        Product.objects
        .select_related("category")
        .filter(
            category=product.category,
            status="active",
        )
        .exclude(pk=product.pk)
        .order_by("-created_at")[:3]
    )

    context = {
        "product": product,
        "related_products": related_products,
    }

    return render(
        request,
        "products/public_detail.html",
        context,
    )
    

def download_price_list_view(request):
    """
    Generate and download the public product price list as Excel.
    Only active products and active categories are included.
    """

    products = (
        Product.objects
        .select_related("category")
        .prefetch_related("price_variations")
        .filter(
            status="active",
            category__status="active",
        )
        .order_by(
            "category__name",
            "name",
        )
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Price List"

    # Excel headers
    headers = [
        "Category",
        "Product",
        "SKU",
        "Variation",
        "Price",
        "Stock",
        "Unit",
        "Stock Status",
    ]

    worksheet.append(headers)

    # Header styling
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="65A832",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # Product + variation data
    for product in products:

        variations = product.price_variations.all()

        for variation in variations:

            worksheet.append([
                product.category.name,
                product.name,
                product.sku or "",
                variation.name,
                variation.price,
                variation.stock,
                variation.get_stock_unit_display(),
                "In Stock" if variation.in_stock else "Out of Stock",
            ])

    # Format Price column
    for cell in worksheet["E"][1:]:
        cell.number_format = '[$₹-en-IN]#,##0.00'

    # Format Stock column
    for cell in worksheet["F"][1:]:
        cell.number_format = '0.00'

    # Freeze header row
    worksheet.freeze_panes = "A2"

    # Add filters
    worksheet.auto_filter.ref = worksheet.dimensions

    # Header row height
    worksheet.row_dimensions[1].height = 25

    # Auto-width columns
    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 3,
            40,
        )

    # Create downloadable response
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    today = date.today().strftime("%Y-%m-%d")

    response["Content-Disposition"] = (
        f'attachment; filename="Mint-Grow-Price-List-{today}.xlsx"'
    )

    workbook.save(response)

    return response


@admin_required
def product_list_view(request):

    products = (
        Product.objects
        .select_related("category")
        .prefetch_related("price_variations")
        .annotate(
            min_price=Min("price_variations__price"),
            max_price=Max("price_variations__price"),
        )
        .order_by("-created_at")
    )

    return render(
        request,
        "products/list.html",
        {
            "products": products,
        },
    )


@admin_required
def product_create_view(request):

    form = ProductForm(
        request.POST or None,
        request.FILES or None,
    )

    price_formset = ProductPriceVariationFormSet(
        request.POST or None,
        prefix="price_variations",
    )

    if request.method == "POST":

        if form.is_valid() and price_formset.is_valid():

            product = form.save()

            price_formset = ProductPriceVariationFormSet(
                request.POST,
                instance=product,
                prefix="price_variations",
            )

            if price_formset.is_valid():
                price_formset.save()

                messages.success(
                    request,
                    f"{product.name} created successfully!",
                )

                return redirect("product_list")

        messages.error(
            request,
            "Please correct the errors below.",
        )

    return render(
        request,
        "products/create.html",
        {
            "form": form,
            "price_formset": price_formset,
        },
    )


@admin_required
@admin_required
def product_update_view(request, id):

    product = get_object_or_404(
        Product,
        id=id,
    )

    old_image = product.image

    form = ProductForm(
        request.POST or None,
        request.FILES or None,
        instance=product,
    )

    price_formset = ProductPriceVariationFormSet(
        request.POST or None,
        instance=product,
        prefix="price_variations",
    )

    if request.method == "POST":

        if form.is_valid() and price_formset.is_valid():

            product = form.save()

            price_formset.save()

            if (
                old_image
                and request.FILES.get("image")
                and old_image != product.image
            ):
                old_image.delete(save=False)

            messages.success(
                request,
                f"{product.name} updated successfully!",
            )

            return redirect("product_list")

        messages.error(
            request,
            "Please correct the errors below.",
        )

    return render(
        request,
        "products/update.html",
        {
            "form": form,
            "price_formset": price_formset,
            "product": product,
        },
    )


@admin_required
@require_POST
def product_toggle_status_view(request, id):

    product = get_object_or_404(
        Product,
        id=id,
    )

    if product.status == "active":

        product.status = "inactive"
        message = "Product deactivated successfully!"

    else:

        product.status = "active"
        message = "Product activated successfully!"

    product.save(
        update_fields=[
            "status",
            "updated_at",
        ]
    )

    messages.success(
        request,
        message,
    )

    return redirect("product_list")


@admin_required
def product_delete_view(request, id):

    product = get_object_or_404(
        Product,
        id=id,
    )

    if request.method == "POST":

        if product.image:
            product.image.delete(save=False)

        product.delete()

        messages.success(
            request,
            f"{product.name} deleted successfully!",
        )

        return redirect("product_list")

    return render(
        request,
        "products/delete.html",
        {
            "product": product,
        },
    )